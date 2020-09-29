import openpyxl
import sqlite3
from datetime import datetime

# open the excel files

# inspections excel sheet
inspectionsWb = openpyxl.load_workbook("inspections.xlsx");
inspectionsSheet = inspectionsWb.active;

# max row / column for inspections.xlsx
inspectionsMaxRow = inspectionsSheet.max_row;
inspectionsMaxColumn = inspectionsSheet.max_column;

# violations excel sheet
violationsWb = openpyxl.load_workbook("violations.xlsx");
violationsSheet = violationsWb.active;

# max row / column for violations.xlsx
violationsMaxRow = violationsSheet.max_row;
violationsMaxColumn = violationsSheet.max_column;

# create the tables for the database
connection = sqlite3.connect("inspections.db");
cursor = connection.cursor();

# drop previous tables / data if they exist (i.e the data may have been updated, or the script may have been previously run - in which case we don't want to duplicate data).
cursor.execute("DROP TABLE IF EXISTS Inspection");
cursor.execute("DROP TABLE IF EXISTS Violation");

# create the table for the inspection data
cursor.execute("""CREATE TABLE Inspection (
				id INTEGER PRIMARY KEY,
				activity_date DATE NOT NULL,
				employee_id VARCHAR(10) NOT NULL,
				facility_address TEXT NOT NULL,
				facility_city VARCHAR(25) NOT NULL,
				facility_id VARCHAR(10) NOT NULL,
				facility_name TEXT NOT NULL,
				facility_state VARCHAR(3),
				facility_zip VARCHAR(15),
				grade CHARACTER(1) NOT NULL,
				owner_id VARCHAR(10) NOT NULL,
				owner_name TEXT NOT NULL,
				pe_description TEXT NOT NULL,
				program_element_pe INTEGER NOT NULL,
				program_name TEXT,
				program_status VARCHAR(10),
				record_id VARCHAR(10) NOT NULL,
				score INTEGER NOT NULL,
				serial_number VARCHAR(10) NOT NULL,
				service_code INTEGER NOT NULL,
				service_description TEXT NOT NULL)""");

# create the table for the violations data
cursor.execute("""CREATE TABLE Violation (
				id INTEGER PRIMARY KEY,
				points INTEGER NOT NULL,
				serial_number VARCHAR(10) NOT NULL,
				violation_code VARCHAR(5) NOT NULL,
				violation_description TEXT,
				violation_status VARCHAR(25))""");


# list containing all the inspection data (each tuple = 1 row of data)
data = [];

# read in the data from inspections.xlsx into the Inspection table
for rowIndex in range(2, inspectionsMaxRow + 1):

	# set list of values to empty each time we access a new row
	rowData = [];

	# read column values for each row
	for columnIndex in range(0, inspectionsMaxColumn):
		rowData.append(inspectionsSheet[chr(ord('A') + columnIndex) + str(rowIndex)].value);
	
	# add the data from the current row to our data list
	data.append(tuple(rowData));

# add all the data to the Inspection table
for values in data:
	cursor.execute("""INSERT INTO Inspection VALUES (NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""", values)


# list containing all the violation data (each tuple = 1 row of data)
data = [];

# read in the data from violations.xlsx into the Violation table
for rowIndex in range(2, violationsMaxRow + 1):

	# set list of values to empty each time we access a new row
	rowData = [];

	# read column values for each row
	for columnIndex in range(0, violationsMaxColumn):
		rowData.append(violationsSheet[chr(ord('A') + columnIndex) + str(rowIndex)].value);
	
	# add the data from the current row to our data list
	data.append(tuple(rowData));

# add all the data to the Violation table
for values in data:
	cursor.execute("""INSERT INTO Violation VALUES (NULL, ?, ?, ?, ?, ?)""", values)

# close the connection and save the state of the database
connection.commit();
connection.close();
